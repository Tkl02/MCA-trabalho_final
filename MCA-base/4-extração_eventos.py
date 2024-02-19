from datetime import datetime
from ics import Calendar
from openpyxl import Workbook
from openpyxl.styles import NamedStyle
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image
import matplotlib.pyplot as plt
import io


def solicitar_mes_ano():
    # Solicita ao usuário o mês desejado
    mes = input("Digite o mês (1-12): ")

    # Valida o mês digitado
    try:
        mes = int(mes)
        if mes < 1 or mes > 12:
            raise ValueError
    except ValueError:
        print("Mês inválido. Digite um número de 1 a 12.")
        return None, None

    # Solicita ao usuário o ano desejado
    ano = input("Digite o ano: ")

    # Valida o ano digitado
    try:
        ano = int(ano)
        if ano < 1:
            raise ValueError
    except ValueError:
        print("Ano inválido. Digite um número maior que 0.")
        return None, None

    return mes, ano


# Solicita ao usuário o mês e o ano desejados
mes, ano = solicitar_mes_ano()

# Verifica se os valores de mês e ano são válidos
if mes is None or ano is None:
    exit()

with open('C:\\Users\\leona\\OneDrive\\Documents\\GitHub\\MCA-trabalho_final\\eventos.ics', 'r', encoding='utf-8') as f:
    conteudo = f.read()


# Cria um objeto Calendar a partir do conteúdo do arquivo
calendario = Calendar(conteudo)

# Filtra os eventos com base no mês e ano escolhidos
eventos_mes_ano = [evento for evento in calendario.events if evento.begin.month == mes and evento.begin.year == ano]

# Verifica se há eventos para o mês e ano escolhidos
if not eventos_mes_ano:
    print("Não há eventos para o mês e ano escolhidos.")
    exit()

# Cria um novo arquivo Excel
workbook = Workbook()
planilha = workbook.active

# Define um estilo para a coluna de datas
data_style = NamedStyle(name="data_style")
data_style.number_format = "DD/MM/YYYY"
planilha.column_dimensions[get_column_letter(1)].width = 12
planilha.column_dimensions['D'].width = 30
# Adiciona os cabeçalhos das colunas
planilha.append(['Data', 'Horário', 'Local', 'Resumo'])

# Extrai os dados dos eventos filtrados e adiciona na planilha
for evento in eventos_mes_ano:
    data = evento.begin.date().strftime("%d/%m/%Y")
    horario = evento.begin.time().strftime("%H:%M:%S")
    local = evento.location if evento.location else "meet"
    resumo = evento.name
    planilha.append([data, horario, local, resumo])

# Aplica o estilo à coluna de datas
for cell in planilha['A']:
    cell.style = data_style

# Salva o arquivo Excel com codificação UTF-8
nome_arquivo = f'dados_reunioes_{datetime.now().strftime("%Y%m%d%H%M%S")}.xlsx'
workbook.save(nome_arquivo)

# Monta o gráfico
dias = list(range(1, 32))
num_reunioes = [len([evento for evento in eventos_mes_ano if evento.begin.day == dia]) for dia in dias]

plt.bar(dias, num_reunioes)
plt.xlabel('Dia')
plt.ylabel('Número de reuniões')
plt.title(f'Reuniões no mês {mes}/{ano}')

# Salva o gráfico em um buffer de imagem
buffer = io.BytesIO()
plt.savefig(buffer, format='png')
buffer.seek(0)

# Cria uma imagem a partir do buffer
imagem = Image(buffer)

# Adiciona a imagem ao arquivo Excel
planilha.add_image(imagem, 'E1')

# Salva o arquivo Excel com a imagem
workbook.save(nome_arquivo)

print(f"Dados das reuniões no mês {mes} e ano {ano} foram salvos no arquivo '{nome_arquivo}'.")


